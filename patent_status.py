import sys, os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from jproperties import Properties
import openpyxl
import datetime
import requests

# load properties from file
configs = Properties()
current_dir = os.path.dirname(os.path.abspath(__file__))
print "current dir: %s" % current_dir
config_file_path = os.path.join(current_dir, 'config.properties')
# debug statement to check if properties file is being loaded correctly
print "properties file path: %s" % config_file_path
with open(config_file_path, 'rb') as read_prop:
    configs.load(read_prop)

# initialize chrome driver
print "driver path from configs: %s" % configs.get("WEBDRIVER_PATH").data
driver_path = os.path.join(current_dir, configs.get("WEBDRIVER_PATH").data)
driver = webdriver.Chrome(driver_path)

patent_status_map = {
    "ACTIVE": "GRANTED",
    "PENDING": "PUBLISHED"
}

patent_xhr_url = str(configs.get("PATENT_XHR_URL").data)


def is_patent_found(patent_no):
    result = requests.get(patent_xhr_url+patent_no).json()
    return not result['error_no_patents_found']


def fetch_patent_status(patent_no):
    driver.get(configs.get("PATENT_URL").data)
    try:
        # wait 3 seconds before looking for elements
        search_box = WebDriverWait(driver, 3).until(
            EC.presence_of_element_located(
                (By.XPATH, "//input[@id='searchInput']"))
        )
        search_box.clear()
        search_box.send_keys(patent_no)
        driver.find_element_by_xpath("//button[@id='searchButton']").send_keys(Keys.RETURN)

        # wait 3 seconds before looking for elements
        status_elem = WebDriverWait(driver, 3).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[@current]"))
        )
        p_status = str(status_elem.text).split('\n')[1].upper()
        p_status = patent_status_map[p_status] if p_status in patent_status_map else p_status
    except Exception as e:
        print "Exception occurred"
	print e
        p_status = "PENDING"
    return p_status


# load the excel file
input_file_loc = os.path.join(current_dir, configs.get("INPUT_FILE_PATH").data)
print "input file location: %s" % input_file_loc
book = openpyxl.load_workbook(input_file_loc)
sheet = book.active
rows = sheet.max_row
patent_no_col = int(configs.get("PATENT_NUMBER_COL").data)
patent_status_col = int(configs.get("PATENT_STATUS_COL").data)
status_ts_col = int(configs.get("STATUS_TS_COL").data)
sheet.cell(row=1, column=status_ts_col).value = "Status Timestamp"

for row in range(2, rows + 1):
    curr_patent_status = str(sheet.cell(row=row, column=patent_status_col).value).upper().strip()
    patent_no = str(sheet.cell(row=row, column=patent_no_col).value)
    sys.stdout.write("\rUpdated Patent  %i of %i" % (row - 1, rows - 1))
    # if status is pending check if patent exists and then check for status
    if patent_no and curr_patent_status and \
            ((curr_patent_status == 'PENDING' and is_patent_found(patent_no))
             or curr_patent_status == 'PUBLISHED'):
        latest_patent_status = fetch_patent_status(patent_no)
        sheet.cell(row=row, column=patent_status_col).value = latest_patent_status
        sheet.cell(row=row, column=status_ts_col).value = datetime.datetime.now()
    sys.stdout.flush()


output_file_path = configs.get("OUTPUT_FILE_PATH").data
print "\noutput file path: %s" % output_file_path
book.save(output_file_path)

driver.quit()
